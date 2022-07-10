import tkinter as tk					
from tkinter import ttk
from tkinter import *
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib
from openpyxl import load_workbook
from PIL import ImageTk
import pandas as pd

root = tk.Tk()
root.title("Education GUI")
root.geometry("600x600")

file = pathlib.Path("StudentData.xlsx")
if file.exists ():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet["A1"] = "Name"
    sheet["B1"] = "PRN"
    sheet["C1"] = "Course"
    sheet["D1"] = "ContactNum"
    sheet["E1"] = "email-id"
    sheet["F1"] = "Address"
    sheet["G1"] = "Skills"
    file.save("StudentData.xlsx")

file1 = pathlib.Path("ResearchData.xlsx")
if file1.exists ():
    pass
else:
    file1 = Workbook()
    sheet1 = file1.active
    sheet1["A1"] = "PRN"
    sheet1["C1"] = "Title"
    sheet1["D1"] = "Publishing Date"
    sheet1["E1"] = "DOI"
    sheet1["B1"] = "Author"
    file1.save("ResearchData.xlsx")

file2 = pathlib.Path("TeachersData.xlsx")
if file2.exists ():
    pass
else:
    file2 = Workbook()
    sheet2 = file2.active
    sheet2["A1"] = "PRN"
    sheet2["B1"] = "Math"
    sheet2["C1"] = "Stats"
    sheet2["D1"] = "Python"
    sheet2["E1"] = "R"
    sheet2["F1"] = "DL"
    sheet2["G1"] = "ML"
    sheet2["H1"] = "Percentage"
    file2.save("TeachersData.xlsx")

file3 = pathlib.Path("AdminData.xlsx")
if file3.exists ():
    pass
else:
    file3 = Workbook()
    sheet3 = file3.active
    sheet3["A1"] = "PRN"
    sheet3["B1"] = "Fees"
    sheet3["C1"] = "PrevMksht"
    sheet3["D1"] = "Aadhar"
    sheet3["E1"] = "LC"
    sheet3["F1"] = "BC"
    file3.save("AdminData.xlsx")
    
tabControl = ttk.Notebook(root, style='lefttab.TNotebook')

myLabel=Label(tabControl, text="STUDENT  DATA", font=("Times New Roman", 25),fg='White', bg="Maroon") #padx=20,pady=20
myLabel.pack()

style = ttk.Style()
style.theme_use('default')
style.configure('lefttab.TNotebook', tabposition='wn', background="Maroon", padding= [30, 35], font=11)
style.configure('TNotebook.Tab', background="Grey", padding= [30, 15], font=11, width=20, height=20)
style.map("TNotebook", background= [("selected", "lavender")])

tab0= ttk.Frame(tabControl)
tab1 = ttk.Frame(tabControl)
tab2 = ttk.Frame(tabControl)
tab3 = ttk.Frame(tabControl)
tab4 = ttk.Frame(tabControl)
#tab5 = ttk.Frame(tabControl)
#tab6 = ttk.Frame(tabControl)

img= PhotoImage(file="C:/Users/desai/Pictures/Screenshots/Screenshot (371).png")
img_labe0= Label(tab0,image=img)
img_label= Label(tab1,image=img)
img_labe2= Label(tab2,image=img)
img_labe3= Label(tab3,image=img)
img_labe4= Label(tab4,image=img)
#img_labe5= Label(tab5,image=img)
#img_labe6= Label(tab6,image=img)

img_labe0.place(x=0, y=0)
img_label.place(x=0, y=0)
img_labe2.place(x=0, y=0)
img_labe3.place(x=0, y=0)
img_labe4.place(x=0, y=0)
#img_labe5.place(x=0, y=0)
#img_labe6.place(x=0, y=0)
#root.config(highlightbackground= "black", highlightthickness=2)

tabControl.add(tab0, text ='Home')
tabControl.add(tab1, text ='Admin')
tabControl.add(tab2, text ='Students')
tabControl.add(tab3, text ='Teachers')
tabControl.add(tab4, text ='Research')
#tabControl.add(tab5, text ='Placement')
#tabControl.add(tab6, text ='Blockchain')
tabControl.pack(expand = 1, fill ="both")

def submit1():
    a =Full_Name.get()
    b = PRN.get()
    c = Course_Name.get()
    f = Contact_number.get()
    g = Email.get()
    h = Address.get()
    p=Skills.get()


    file = openpyxl.load_workbook("StudentData.xlsx")
    sheet = file.active
    
    sheet.cell(column = 1, row = sheet.max_row+1, value = a)
    sheet.cell(column = 2, row = sheet.max_row, value = b)
    sheet.cell(column = 3, row = sheet.max_row, value = c)
    sheet.cell(column = 4, row = sheet.max_row, value = f)
    sheet.cell(column = 5, row = sheet.max_row, value = g)
    sheet.cell(column = 6, row = sheet.max_row, value = h)
    sheet.cell(column = 7, row = sheet.max_row, value = p)

    file.save("StudentData.xlsx")

xlfile = pd.read_excel ('C:/Users/desai/AppData/Local/Programs/Python/Python310/StudentData.xlsx', 'Sheet')
xlfile.to_csv('C:/Users/desai/AppData/Local/Programs/Python/Python310/StudentData.csv', index = False)

def submit2():
    aa=PRN1.get()
    s=Auth.get()
    j = Research_Title.get()
    k = Date.get()
    l = DOI.get()

    file1 = openpyxl.load_workbook("ResearchData.xlsx")
    sheet1 = file1.active
    
    sheet1.cell(column = 1, row = sheet1.max_row+1, value = aa)
    sheet1.cell(column = 2, row = sheet1.max_row, value = s)
    sheet1.cell(column = 3, row = sheet1.max_row, value = j)
    sheet1.cell(column = 4, row = sheet1.max_row, value = k)
    sheet1.cell(column = 5, row = sheet1.max_row, value = l)
        
    file1.save("ResearchData.xlsx")

xlfile1 = pd.read_excel ('C:/Users/desai/AppData/Local/Programs/Python/Python310/ResearchData.xlsx', 'Sheet') 
xlfile1.to_csv('C:/Users/desai/AppData/Local/Programs/Python/Python310/ResearchData.csv', index = False)

def submit3():
    bb=PRN2.get()
    q = var.get()
    z=Math.get()
    x=Stats.get()
    y=Python.get()
    w=R.get()
    v=DL.get()
    u=ML.get()
    
    file2 = openpyxl.load_workbook("TeachersData.xlsx")
    sheet2 = file2.active
    
    sheet2.cell(column = 1, row = sheet2.max_row+1, value = bb)
    sheet2.cell(column = 2, row = sheet2.max_row, value = z)
    sheet2.cell(column = 3, row = sheet2.max_row, value = x)
    sheet2.cell(column = 4, row = sheet2.max_row, value = y)
    sheet2.cell(column = 5, row = sheet2.max_row, value = w)
    sheet2.cell(column = 6, row = sheet2.max_row, value = v)
    sheet2.cell(column = 7, row = sheet2.max_row, value = u)
    sheet2.cell(column = 8, row = sheet2.max_row, value = q)
        
    file2.save("TeachersData.xlsx")

xlfile2 = pd.read_excel ('C:/Users/desai/AppData/Local/Programs/Python/Python310/TeachersData.xlsx', 'Sheet') 
xlfile2.to_csv('C:/Users/desai/AppData/Local/Programs/Python/Python310/TeachersData.csv', index = False)

def submit4():
    cc=PRN3.get()
    r = var2.get()
    s = var3.get()
    t = var4.get()
    u = var5.get()
    v = var6.get()

    file3 = openpyxl.load_workbook("AdminData.xlsx")
    sheet3 = file3.active
    
    sheet3.cell(column = 1, row = sheet3.max_row+1, value = cc)
    sheet3.cell(column = 2, row = sheet3.max_row, value = r)
    sheet3.cell(column = 3, row = sheet3.max_row, value = s)
    sheet3.cell(column = 4, row = sheet3.max_row, value = t)
    sheet3.cell(column = 5, row = sheet3.max_row, value = u)
    sheet3.cell(column = 6, row = sheet3.max_row, value = v)
        
    file3.save("AdminData.xlsx")

xlfile1 = pd.read_excel ('C:/Users/desai/AppData/Local/Programs/Python/Python310/AdminData.xlsx', 'Sheet') 
xlfile1.to_csv('C:/Users/desai/AppData/Local/Programs/Python/Python310/AdminData.csv', index = False)

col_count, row_count = tab2.grid_size()

for col in range(col_count):
    tab2.grid_columnconfigure(col, minsize=50)

for row in range(row_count):
    tab2.grid_rowconfigure(row, minsize=50)

Label(tab2, text = "Full Name:").grid(row=6, column=0, padx= 15, pady= 15,  sticky='w')
Label(tab2, text = "PRN:").grid(row=8, column=0, padx= 15, pady= 15,  sticky='w')
Label(tab2, text = "Course Name:").grid(row=10, column=0, padx= 15, pady= 15,  sticky='w')
Label(tab2, text = "Contact_number:").grid(row=12, column=0, padx= 15, pady= 15,  sticky='w')
Label(tab2, text = "Personal email-id:").grid(row=14, column=0, padx= 15, pady= 15,  sticky='w')
Label(tab2, text = "Permanent Residential Address:").grid(row=16, column=0, padx= 15, pady= 15,  sticky='w')
Label(tab2, text = "Top 5 Skills:").grid(row=18, column=0, padx= 18, pady= 15,  sticky='w')

Label(tab4, text = "Title of Research Paper:").grid(row=8, column=0, padx= 15, pady= 15,  sticky='w')
Label(tab4, text = "Date of Publishing:").grid(row=10, column=0, padx= 15, pady= 15,  sticky='w')
Label(tab4, text = "DOI:").grid(row=12, column=0, padx= 15, pady= 15,  sticky='w')
Label(tab4, text = "Author:").grid(row=14, column=0, padx= 15, pady= 15,  sticky='w')
Label(tab4, text = "PRN:").grid(row=6, column=0, padx= 15, pady= 15,  sticky='w')

Label(tab3, text = "Maths Marks:").grid(row=8, column=0, padx= 15, pady= 15,  sticky='w')
Label(tab3, text = "Stats Marks:").grid(row=10, column=0, padx= 15, pady= 15,  sticky='w')
Label(tab3, text = "Python Marks:").grid(row=12, column=0, padx= 15, pady= 15,  sticky='w')
Label(tab3, text = "R Programming Marks:").grid(row=14, column=0, padx= 15, pady= 15,  sticky='w')
Label(tab3, text = "Deep Learning Marks:").grid(row=16, column=0, padx= 15, pady= 15,  sticky='w')
Label(tab3, text = "Machine Learning Marks:").grid(row=18, column=0, padx= 15, pady= 15,  sticky='w')
Label(tab3, text = "PRN:").grid(row=6, column=0, padx= 15, pady= 15,  sticky='w')

Label(tab1, text = "PRN:").grid(row=6, column=0, padx= 15, pady= 15,  sticky='w')

Full_Name = Entry(tab2, width=45)
Full_Name.grid(row=6, column=1, padx= 15, pady= 15,  sticky='w')


PRN = Entry(tab2)
PRN.grid(row=8, column=1,padx= 15, pady= 15,  sticky='w')

PRN3 = Entry(tab1)
PRN3.grid(row=6, column=1,padx= 15, pady= 15,  sticky='w')

PRN2 = Entry(tab3)
PRN2.grid(row=6, column=1,padx= 15, pady= 15,  sticky='w')

PRN1 = Entry(tab4)
PRN1.grid(row=6, column=1,padx= 15, pady= 15,  sticky='w')


Course_Name = Entry(tab2, width=50)
Course_Name.grid(row=10, column=1, padx= 15, pady= 15,  sticky='w')

Contact_number = Entry(tab2)
Contact_number.grid(row=12, column=1, padx= 15, pady= 15,  sticky='w')

Email = Entry(tab2, width=50)
Email.grid(row=14, column=1, padx= 15, pady= 15,  sticky='w')

Address = Entry(tab2, width=60)
Address.grid(row=16, column=1,padx= 15, pady= 15,  sticky='w')

Skills= Entry(tab2, width=60)
Skills.grid(row=18, column=1, padx= 15, pady= 15,  sticky='w')

Research_Title = Entry(tab4, width=60)
Research_Title.grid(row=8, column=1, padx= 15, pady= 15,  sticky='w')

Date = Entry(tab4)
Date.grid(row=10, column=1, padx= 15, pady= 15,  sticky='w')

DOI = Entry(tab4, width=45)
DOI.grid(row=12, column=1, padx= 15, pady= 15,  sticky='w')

Auth = Entry(tab4, width=45)
Auth.grid(row=14, column=1, padx= 15, pady= 15,  sticky='w')

Math = Entry(tab3)
Math.grid(row=8, column=1, padx= 15, pady= 15,  sticky='w')

Stats = Entry(tab3)
Stats.grid(row=10, column=1, padx= 15, pady= 15,  sticky='w')

Python = Entry(tab3)
Python.grid(row=12, column=1, padx= 15, pady= 15,  sticky='w')

R = Entry(tab3)
R.grid(row=14, column=1, padx= 15, pady= 15,  sticky='w')

DL = Entry(tab3)
DL.grid(row=16, column=1, padx= 15, pady= 15,  sticky='w')

ML = Entry(tab3)
ML.grid(row=18, column=1, padx= 15, pady= 15,  sticky='w')


label_2 = Label(tab1, text = "Fee Payment Status",width = 20, font = ("bold",10))
label_2.grid(row=8, column=0, padx= 15, pady= 15,  sticky='w')
var2 = IntVar()
Radiobutton(tab1, text = "Paid", padx = 5, variable = var2, value = 1).grid(row=8, column=1, padx= 15, pady= 15)
Radiobutton(tab1, text = "Pending", padx = 5, variable =  var2, value = 2).grid(row=8, column=2, padx= 15, pady= 15)

label_3 = Label(tab1, text = "Previous Marksheet Copy",width = 20, font = ("bold",10))
label_3.grid(row=10, column=0, padx= 15, pady= 15,  sticky='w')
var3 = IntVar()
Radiobutton(tab1, text = "Submitted", padx = 5, variable = var3, value = 1).grid(row=10, column=1, padx= 15, pady= 15)
Radiobutton(tab1, text = "Not Submitted", padx = 5, variable =  var3, value = 2).grid(row=10, column=2, padx= 15, pady= 15)

label_4 = Label(tab1, text = "Aadhar Card Copy",width = 20, font = ("bold",10))
label_4.grid(row=12, column=0, padx= 15, pady= 15,  sticky='w')
var4 = IntVar()
Radiobutton(tab1, text = "Submitted", padx = 5, variable = var4, value = 1).grid(row=12, column=1, padx= 15, pady= 15)
Radiobutton(tab1, text = "Not Submitted", padx = 5, variable =  var4, value = 2).grid(row=12, column=2, padx= 15, pady= 15)

label_5 = Label(tab1, text = "Leaving Certificate",width = 20, font = ("bold",10))
label_5.grid(row=14, column=0, padx= 15, pady= 15,  sticky='w')
var5 = IntVar()
Radiobutton(tab1, text = "Submitted", padx = 5, variable = var5, value = 1).grid(row=14, column=1, padx= 15, pady= 15)
Radiobutton(tab1, text = "Not Submitted", padx = 5, variable =  var5, value = 2).grid(row=14, column=2, padx= 15, pady= 15)

label_6 = Label(tab1, text = "Birth Certificate Copy",width = 20, font = ("bold",10))
label_6.grid(row=16, column=0, padx= 15, pady= 15,  sticky='w')
var6 = IntVar()
Radiobutton(tab1, text = "Submitted", padx = 5, variable = var6, value = 1).grid(row=16, column=1, padx= 15, pady= 15)
Radiobutton(tab1, text = "Not Submitted", padx = 5, variable =  var6, value = 2).grid(row=16, column=2, padx= 15, pady= 15)

#Percentage SLIDER
label_8 = Label(tab3, text = "Select Percentage value: ",width = 20, font = ("bold",10))
label_8.grid(row=20, column=0, padx= 15, pady= 15)

def sel():
   selection = ":" + str(var.get())
   label_7.config(text = selection)

var = DoubleVar()
scale = Scale( tab3, variable = var , orient='horizontal')
scale.grid(row=20, column=1, padx= 15, pady= 15)

button = Button(tab3, text="Get Percentage value", command=sel)
button.grid(row=20, column=2, padx= 15, pady= 15)

label_7 = Label(tab3)
#label.pack()
label_7.grid(row=20, column=3, padx= 15, pady= 15)



Button(tab2,text = "Submit Response", command = submit1).grid(row=20, column=1, padx= 15, pady= 15)
Button(tab4,text = "Submit Response", command = submit2).grid(row=16, column=1, padx= 15, pady= 15)
Button(tab3,text = "Submit Response", command = submit3).grid(row=22, column=1, padx= 15, pady= 15)
Button(tab1,text = "Submit Response", command = submit4).grid(row=18, column=1, padx= 15, pady= 15)











'''

#BLOCKCHAIN

#Importing libraries & data
import datetime as dt #For timestamp
import hashlib as hl  #Calculating the hash in order to add digital fingerprints to the blocks
from flask import Flask, jsonify # Flask is for creating the web app and jsonify is for displaying the blockchain
import json # To store data in our blockchain
import csv

# Function to convert a CSV to JSON
def make_json(csvFilePath, jsonFilePath):
    data = {}
     
    # Open a csv reader called DictReader
    with open(csvFilePath, encoding='utf-8') as csvf:
        csvReader = csv.DictReader(csvf)
         
        # Convert each row into a dictionary
        # and add it to data
        for rows in csvReader:
             
            # Assuming a column named 'PRN' to  be the primary key
            key = rows['PRN']
            data[key] = rows
 
    # Open a json writer, and use the json.dumps() function to dump data
    with open(jsonFilePath, 'w', encoding='utf-8') as jsonf:
        jsonf.write(json.dumps(data, indent=4))

        
 
# Decide the two file paths according to your computer system
csvFilePath1 = r'C:/Users/desai/AppData/Local/Programs/Python/Python310/StudentData.csv'
jsonFilePath1 = r'C:/Users/desai/AppData/Local/Programs/Python/Python310/StudentData.json'
# Call the make_json function
make_json(csvFilePath1, jsonFilePath1)

with open('C:/Users/desai/AppData/Local/Programs/Python/Python310/StudentData.json') as json_file:
    json_data = json.load(json_file)
#print(json_data)

csvFilePath2 = r'C:/Users/desai/AppData/Local/Programs/Python/Python310/TeachersData.csv'
jsonFilePath2 = r'C:/Users/desai/AppData/Local/Programs/Python/Python310/TeachersData.json'
make_json(csvFilePath2, jsonFilePath2)
with open('C:/Users/desai/AppData/Local/Programs/Python/Python310/TeachersData.json') as json_file2:
    json_data2 = json.load(json_file2)

csvFilePath3 = r'C:/Users/desai/AppData/Local/Programs/Python/Python310/ResearchData.csv'
jsonFilePath3 = r'C:/Users/desai/AppData/Local/Programs/Python/Python310/ResearchData.json'
make_json(csvFilePath3, jsonFilePath3)
with open('C:/Users/desai/AppData/Local/Programs/Python/Python310/ResearchData.json') as json_file3:
    json_data3 = json.load(json_file3)

csvFilePath4 = r'C:/Users/desai/AppData/Local/Programs/Python/Python310/AdminData.csv'
jsonFilePath4 = r'C:/Users/desai/AppData/Local/Programs/Python/Python310/AdminData.json'
make_json(csvFilePath4, jsonFilePath4)
with open('C:/Users/desai/AppData/Local/Programs/Python/Python310/AdminData.json') as json_file4:
    json_data4 = json.load(json_file4)



class Block:
    def __init__(self, id, timestamp, proof, data, parent_hash):
        self.id=id
        self.timestamp=timestamp
        self.data=data
        self.parent_hash=parent_hash
        self.proof=proof

class Blockchain:
    def __init__(self):
        a=Block(id=1, timestamp= str(dt.datetime.now()), data= 'gen_data',parent_hash='0', proof=1)
        genesis= a.__dict__
        self.chain= [genesis]
        
    def add_block(self, data):    
        id=self.chain[-1]['id'] + 1
        p=self.chain[-1]
        parent_hash= hl.sha256(json.dumps(p).encode()).hexdigest()
        proof= Blockchain().POW(self.chain[-1]['proof'])
        b=Block(id, str(dt.datetime.now()), proof, data, parent_hash)
        block=b.__dict__
        self.chain.append(block) #To add the new block to the chain
        return block #, proof
    
    def POW(self, parent_pf):
        new_pf= 1        #new proof
        check_pf= False  #check proof
        
        while check_pf is False:
            #hash operation
            hash_op= hl.sha256(str(new_pf**2 - parent_pf**2).encode()).hexdigest()
            
            if hash_op[:5]== '00000':
                check_pf= True
            else:
                new_pf+=1
        return new_pf
    
    def hashGenerator(self, block):
        return hl.sha256(json.dumps(block).encode()).hexdigest() #'utf-8' #encode, create hash, convert to hexadecimal   
    
    def check_chain_validity(self, chain):
        parent_block= chain[0]
        block_index= 1
        
        #Checking parent hash of every block
        while block_index < len(chain):
            block= chain[block_index]
            if block['parent_hash']!=self.hashGenerator(parent_block):
                return False
            
            #Checking POW for every block
            parent_pf= parent_block['proof']
            proof= block['proof']
            hash_op= hl.sha256(str(proof**2 - parent_pf**2).encode()).hexdigest()
            
            if hash_op[:5]!= '00000':
                return False
            #Now, 2nd block will be parent block & new block will be the 3rd block (Hence, we're incrementing):
            parent_block= block
            block_index+=1
        
        return True
            
            
bc= Blockchain()

#bc.add_block(json_data1)
#bc.add_block(json_data2)
#print(bc.chain)

#for block in bc.chain:
 #   print(block)


def valid():
    valid = bc.check_chain_validity(bc.chain)
     
    if valid:
        response = 'The Blockchain is valid.'
    else:
        response = 'The Blockchain is not valid.'
    return response

v1=Scrollbar(tab6, orient='vertical')
v2=Scrollbar(tab6, orient='vertical')
v3=Scrollbar(tab6, orient='vertical')
v4=Scrollbar(tab6, orient='vertical')

v1.grid(column=2, row=2, rowspan=2,  sticky=N+S+W)
v2.grid(column=2, row=4, rowspan=2,  sticky=N+S+W)
v3.grid(column=2, row=6, rowspan=2,  sticky=N+S+W)
v4.grid(column=2, row=8, rowspan=2,  sticky=N+S+W)

def display_results():
    show_up = bc.add_block(json_data)
    text1 = tk.Text(tab6, height=8, width=100, yscrollcommand=v2.set)
    v2.config(command=text1.yview)
    text1.grid(column=2, row=4, padx= 11, pady= 11,  sticky='w')
    text1.insert(tk.END, show_up)
    text1["state"] = DISABLED
    
Btn=tk.Button(tab6, text="Add Student Data Block: ", command= display_results)
Btn.grid(row=4, column=1, padx= 15, pady= 15,  sticky='w')




def display_results2():
    show_up2 = bc.add_block(json_data4)
    text2 = tk.Text(tab6, height=8, width=100, yscrollcommand=v1.set)
    v1.config(command=text2.yview)
    text2.grid(column=2, row=2, padx= 11, pady= 11,  sticky='w')
    text2.insert(tk.END, show_up2)
    text2["state"] = DISABLED
    
Btn2=tk.Button(tab6, text="Add Admin Data Block: ", command= display_results2)
Btn2.grid(row=2, column=1, padx= 15, pady= 15,  sticky='w')

def display_results3():
    show_up3 = bc.add_block(json_data2)
    text3 = tk.Text(tab6, height=8, width=100, yscrollcommand=v3.set)
    v3.config(command=text3.yview)
    text3.grid(column=2, row=6, padx= 11, pady= 11,  sticky='w')
    text3.insert(tk.END, show_up3)
    text3["state"] = DISABLED
    
Btn3=tk.Button(tab6, text="Add Teachers Data Block: ", command= display_results3)
Btn3.grid(row=6, column=1, padx= 15, pady= 15,  sticky='w')

def display_results4():
    show_up4 = bc.add_block(json_data3)
    text4 = tk.Text(tab6, height=8, width=100, yscrollcommand=v4.set)
    v4.config(command=text4.yview)
    text4.grid(column=2, row=8, padx= 11, pady= 11,  sticky='w')
    text4.insert(tk.END, show_up4)
    text4["state"] = DISABLED
    
Btn4=tk.Button(tab6, text="Add Research Data Block: ", command= display_results4)
Btn4.grid(row=8, column=1, padx= 15, pady= 15,  sticky='w')





def display_results1():    
    show_up1 = valid()
    text = tk.Text(tab6, height=3, width=80)
    text.grid(column=2, row=10, padx= 10, pady= 10,  sticky='w')
    text.insert(tk.END, show_up1)
    

Btn1=tk.Button(tab6, text="Print Blockchain Validity: ", command= display_results1)
Btn1.grid(row=10, column=1, padx= 15, pady= 15,  sticky='w')


'''

root.mainloop()
